# from cv2 import VideoCapture
# from cv2 import imwrite
import os
import openpyxl
from openpyxl import load_workbook
import re
import math
# import xlrd
# import xlwt
# _*_ coding:utf-8 _*_
if __name__ == '__main__':

        predict_dir = 'D:/Desktop/SZX'
        excel_dir = 'D:/Desktop/SZX/'
        predict_name = os.listdir(predict_dir)
        x0 = 0
        y0 = 0
        for index in range(len(predict_name)):
            # predict_name下对应的img文件名列表
            #labels_dir = predict_dir + '/' + predict_name[index] + '/labels'
            labels_dir = predict_dir + '/' + '/labels'
            img_name = os.listdir(labels_dir)
            # img文件路径+名字
            img_dir = [labels_dir + '/' + str(i) for i in img_name]

            # 判断路径是否存在，存在则返回true
            isExists = os.path.exists(excel_dir)
            if not isExists:
                os.mkdir(excel_dir)

            excel_filename = excel_dir + predict_name[index] + '.xlsx'
            if not os.path.exists(excel_filename):
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = "原始数据"
                sheet.append(["pre_name", "img_name", "序号", "类型", "X0", "Y0", "WIDTH", "HEIGHT", "置信度", "线长", "x距离", "y距离", "欧拉"])
            else:
                # 打开工作簿
                wb = load_workbook(excel_filename)
                sheet = wb.active

            for img_index in range(0, len(img_name), 1):

                # 依次读取每行
                fp = open(img_dir[img_index], 'r')
                for line in fp.readlines():
                    # predict_name img_name number

                    rows = [predict_name[index], img_name[img_index]]
                    img_num = int(re.split('[_ .]', img_name[img_index])[1])
                    # print(img_num)
                    rows.extend([img_num])
                    # 加txt数据
                    line = line.strip()  # get rid of trailing line break character(s)
                    txt_row = re.split(r" +", line)[0:6]
                    txt_row_float = list(map(float, txt_row))
                    if txt_row_float[0] == 0:
                        rows.extend(txt_row_float)
                    # 加线长
                        confidence = float(txt_row[0])
                        width = float(txt_row[3])
                        height = float(txt_row[4])
                        wire_length = math.sqrt(pow(width, 2) + pow(height, 2))
                        rows.extend([wire_length])
                        x1 = float(txt_row[1]) - x0
                        x0 = float(txt_row[1])
                        rows.extend([x1])
                        y1 = float(txt_row[2]) - y0
                        y0 = float(txt_row[2])
                        rows.extend([y1])
                        #if y1 != 0 and 0.47 > y0 > 0.4 :
                        sheet.append(rows)
                        rows.clear()

                fp.close()

            #target = wb.copy_worksheet(sheet,"数据处理")
            # sheet_name = "Sheet Copy"
            # 保存文件
            wb.save(excel_filename)

            print(predict_name[index])
